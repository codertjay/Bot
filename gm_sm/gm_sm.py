import html
import json

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl

BASE_URL = "https://tzdubrovnik.hr/news/gm_smjestaj/index.html"
from time import sleep

ALL_LINKS = []


class GmSM:
    def __init__(self, teardown=True):

        s = Service(ChromeDriverManager().install())
        self.options = webdriver.ChromeOptions()
        self.options.add_argument('headless')
        self.teardown = teardown
        # keep chrome open
        self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option(
            "excludeSwitches",
            ['enable-logging'])
        self.driver = webdriver.Chrome(
            options=self.options,
            service=s)
        self.collection = []
        self.driver.implicitly_wait(50)
        super(GmSM, self).__init__()

    def __enter__(self):
        self.driver.get(BASE_URL)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.driver.quit()

    def land_first_page(self):
        self.driver.get(BASE_URL)

    def get_emails_on_subs(self):
        for item in range(70):
            if item > 70:
                break
            self.driver.get(f"https://tzdubrovnik.hr/news/gm_smjestaj/page{item + 1}.html")
            try:
                elements = self.driver.find_elements(
                    by=By.CSS_SELECTOR, value='a[class="tz_ni_li_val text_ellip_1"]')
                for element in elements:
                    email = element.get_attribute("href").replace("mailto:",
                                                                  "")
                    if not email.startswith("http"):
                        self.collection.append(
                            {
                                "email": email,
                            }
                        )
                        print(email)

                        with open('GmSM.json', 'w') as f:
                            f.write(json.dumps(self.collection))
            except Exception as a:
                pass

    def convert_to_xl(self):
        file = open('GmSM.json', 'r')
        read_file = file.read()
        collection = json.loads(read_file)
        # Create a new Excel file
        workbook = openpyxl.Workbook()

        # Get the active sheet
        worksheet = workbook.active

        # Add the headers to the sheet
        headers = list(collection[0].keys())  # Get the keys of the first dictionary in the list
        for col, header in enumerate(headers):
            worksheet.cell(row=1, column=col + 1).value = header

        # Add the collection to the sheet
        for row, item in enumerate(collection):
            for col, value in enumerate(item.values()):
                worksheet.cell(row=row + 2, column=col + 1).value = value

        # Save the file
        workbook.save("GmSM.xlsx")
