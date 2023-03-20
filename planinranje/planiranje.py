import html
import json

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl

BASE_URL = "https://www.infozagreb.hr/planiranje-putovanja/smjestaj/privatni-smjestaj/apartman"


class Planiranje:
    def __init__(self, teardown=True):

        s = Service(ChromeDriverManager().install())
        self.options = webdriver.ChromeOptions()
        # self.options.add_argument('headless')
        self.teardown = teardown
        # keep chrome open
        # self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option(
            "excludeSwitches",
            ['enable-logging'])
        self.driver = webdriver.Chrome(
            options=self.options,
            service=s)
        self.driver.implicitly_wait(50)
        self.collection = []
        super(Planiranje, self).__init__()

    def __enter__(self):
        self.driver.get(BASE_URL)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.driver.quit()

    def land_first_page(self):
        self.driver.get(BASE_URL)

    def get_details(self):
        templates = self.driver.find_element(
            by=By.CSS_SELECTOR, value='div[class="gridView medium"]'
        ).find_elements(by=By.TAG_NAME, value="li")
        for element in templates:
            try:
                name = html.escape(element.find_element(
                    by=By.CSS_SELECTOR, value='div[class="text"]').find_element(
                    by=By.TAG_NAME, value='a').find_element(
                    by=By.TAG_NAME, value='h2').get_attribute('innerHTML').strip())
            except:
                name = None
            try:
                location = html.escape(element.find_element(
                    by=By.CSS_SELECTOR, value='div[class="text"]').find_element(
                    by=By.TAG_NAME, value='a').find_element(
                    by=By.TAG_NAME, value='p').get_attribute('innerHTML').strip())
            except:
                location = None
            try:
                email = element.find_element(
                    by=By.CSS_SELECTOR, value='div[class="text"]').find_element(
                    by=By.CSS_SELECTOR, value='a[class="sprite gridMail"]').get_attribute("href").strip()
                if email:
                    email = email.replace("mailto:", "")
            except:
                email = None

            self.collection.append(
                {
                    "name": name,
                    "location": location,
                    "email": email,
                }
            )
            with open('Planiranje.json', 'w') as f:
                f.write(json.dumps(self.collection))

    def convert_to_xl(self):
        file = open('Planiranje.json', 'r')
        read_file = file.read()
        collection = json.loads(read_file)
        # Create a new Excel file
        workbook = openpyxl.Workbook()

        # Get the active sheet
        worksheet = workbook.active

        # Add the headers to the sheet
        headers = list(collection[0].keys())  # Get the keys of the first dictionary in the list
        for col, header in enumerate(headers):
            worksheet.cell(row=1, column=col + 1).value = header

        # Add the collection to the sheet
        for row, item in enumerate(collection):
            for col, value in enumerate(item.values()):
                worksheet.cell(row=row + 2, column=col + 1).value = value

        # Save the file
        workbook.save("data.xlsx")
