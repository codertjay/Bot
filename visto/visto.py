import html
import json

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl

BASE_URL = "https://www.visitopatija.com/en"
from time import sleep

ALL_LINKS = ['https://www.visitopatija.com/en/dreamtime-p2089',
             'https://www.visitopatija.com/en/villa-de-zamagna-apartment-camelia-p1883',
             'https://www.visitopatija.com/en/apartments-ivan-curcic-p1861',
             'https://www.visitopatija.com/en/apartment-bucic-milovan-2-p1849',
             'https://www.visitopatija.com/en/apartments-anamarija-p1006',
             'https://www.visitopatija.com/en/apartment-biser-p1938',
             'https://www.visitopatija.com/en/villa-de-zamagna-apartment-magnolia-p1885',
             'https://www.visitopatija.com/en/apartment-kalina-hills-p1623',
             'https://www.visitopatija.com/en/apartment-beti-p1934',
             'https://www.visitopatija.com/en/sunny-apartment-anamaria-p954', 'https://www.visitopatija.com/en/-p2050',
             'https://www.visitopatija.com/en/apartments-lavanda-p1604',
             'https://www.visitopatija.com/en/apartment-popek-p1893',
             'https://www.visitopatija.com/en/apartment-teo-p1951',
             'https://www.visitopatija.com/en/apartment-oleandri-p1571',
             'https://www.visitopatija.com/en/apartment-hortensia-2-p1891',
             'https://www.visitopatija.com/en/house-for-rest-jakupovic-p1262',
             'https://www.visitopatija.com/en/apartment-gloria-p1124',
             'https://www.visitopatija.com/en/apartment-ascic-gudeljevic-anica-2-p999',
             'https://www.visitopatija.com/en/apartments-grozic-1-p1106',
             'https://www.visitopatija.com/en/villa-st-maria-p1878',
             'https://www.visitopatija.com/en/apartment-liana-p1925',
             'https://www.visitopatija.com/en/apartment-silvi-p1631', 'https://www.visitopatija.com/en/des-4u-p1942',
             'https://www.visitopatija.com/en/apartment-opatija-residenz-p2092',
             'https://www.visitopatija.com/en/apartments-agata-p1856',
             'https://www.visitopatija.com/en/apartments-lucija-p1243',
             'https://www.visitopatija.com/en/apartment-glasnovic-2-p1846',
             'https://www.visitopatija.com/en/villa-antonio-p1626',
             'https://www.visitopatija.com/en/apartment-opatija-2-p1544',
             'https://www.visitopatija.com/en/apartment-tennis-p1933',
             'https://www.visitopatija.com/en/petra-hills-p1415',
             'https://www.visitopatija.com/en/apartmans-vladislav-vivoda-p1780',
             'https://www.visitopatija.com/en/casa-salita-garibaldi-p1943',
             'https://www.visitopatija.com/en/apartment-castel-san-pietro-p1836',
             'https://www.visitopatija.com/en/apartment-milovanovic-marko-primorska-21-ika-p1481',
             'https://www.visitopatija.com/en/villa-de-zamagna-apartment-lungo-mare-p1886',
             'https://www.visitopatija.com/en/apartment-volosko-p2047',
             'https://www.visitopatija.com/en/apartment-marija-p1453',
             'https://www.visitopatija.com/en/villa-de-zamagna-apartment-sv-jakov-p1888',
             'https://www.visitopatija.com/en/apartment-vila-sokol-p1900',
             'https://www.visitopatija.com/en/gherliza-zeljka-p1190',
             'https://www.visitopatija.com/en/apartment-iris-p1851',
             'https://www.visitopatija.com/en/apartment-ana-p1144',
             'https://www.visitopatija.com/en/apartment-ola-ii-p2028',
             'https://www.visitopatija.com/en/apartment-tomic-p923',
             'https://www.visitopatija.com/en/apartment-vili-681-p1275',
             'https://www.visitopatija.com/en/apartment-liana-p1926',
             'https://www.visitopatija.com/en/apartment-iva-opatija-center-p1860',
             'https://www.visitopatija.com/en/-p2091', 'https://www.visitopatija.com/en/apartment-ivancic-p1469',
             'https://www.visitopatija.com/en/apartments-vivoda-lina-p1377',
             'https://www.visitopatija.com/en/kvarner-star-p2062', 'https://www.visitopatija.com/en/villa-bella-p1122',
             'https://www.visitopatija.com/en/mirko-bulic-p1495',
             'https://www.visitopatija.com/en/apartment-anicic-p1147',
             'https://www.visitopatija.com/en/apartment-nostalgie-p1895',
             'https://www.visitopatija.com/en/apartment-inga-p1232',
             'https://www.visitopatija.com/en/apartment-rubin-p1602',
             'https://www.visitopatija.com/en/apartment-moho-p1833',
             'https://www.visitopatija.com/en/apartment-evici-p1491',
             'https://www.visitopatija.com/en/apartments-anicic-mauro-p2063',
             'https://www.visitopatija.com/en/apartment-zuzzi-p1718',
             'https://www.visitopatija.com/en/apartment-oli-p933',
             'https://www.visitopatija.com/en/apartmani-glasnovic-p992',
             'https://www.visitopatija.com/en/buona-vista-apartment-opatija-p1546',
             'https://www.visitopatija.com/en/smart-4-sports-apartment-p1896',
             'https://www.visitopatija.com/en/apartmant-opatija-beach-p2046',
             'https://www.visitopatija.com/en/apartments-amorino-p1408',
             'https://www.visitopatija.com/en/apartment-ascic-gudeljevic-anica-1-p964',
             'https://www.visitopatija.com/en/apartman-iva-p1545',
             'https://www.visitopatija.com/en/urban-villa-apartment-p2064',
             'https://www.visitopatija.com/en/apartment-p2040',
             'https://www.visitopatija.com/en/apartment-marija-p2060',
             'https://www.visitopatija.com/en/apartment-glasnovic-3-p1847',
             'https://www.visitopatija.com/en/simunic-damir-p1668',
             'https://www.visitopatija.com/en/apartment-gloria-2-p1637',
             'https://www.visitopatija.com/en/apartments-juraga-villa-salone-p1110',
             'https://www.visitopatija.com/en/villa-krasa-p1148', 'https://www.visitopatija.com/en/apartment-ola-p1930',
             'https://www.visitopatija.com/en/-p2085', 'https://www.visitopatija.com/en/apartment-djolonga-mirko-p1496',
             'https://www.visitopatija.com/en/apartment-hortensia-1-p1233',
             'https://www.visitopatija.com/en/apartmani-garden-opatija-p991',
             'https://www.visitopatija.com/en/apartment-san-pietro-p1146',
             'https://www.visitopatija.com/en/apartment-anita-p969',
             'https://www.visitopatija.com/en/panoramico-apartment-opatija-p1931',
             'https://www.visitopatija.com/en/apartment-oliva-p1876',
             'https://www.visitopatija.com/en/apartments-amorino-p1588', 'https://www.visitopatija.com/en/-p2087',
             'https://www.visitopatija.com/en/apartman-albrecht-p2033',
             'https://www.visitopatija.com/en/apartment-damir-kramer-p1082',
             'https://www.visitopatija.com/en/apartment-gea-p1315',
             'https://www.visitopatija.com/en/apartments-grozic-2-p1456',
             'https://www.visitopatija.com/en/villa-antonija-p1889',
             'https://www.visitopatija.com/en/apartment-rosenberger-p2078',
             'https://www.visitopatija.com/en/apartment-gluhak-p1719',
             'https://www.visitopatija.com/en/sun-and-moon-apartment-p1879', 'https://www.visitopatija.com/en/-p2049',
             'https://www.visitopatija.com/en/apartment-jasmina-p1783',
             'https://www.visitopatija.com/en/deluxe-apartments-opatija-p1899',
             'https://www.visitopatija.com/en/apartment-o-la-la-p1869',
             'https://www.visitopatija.com/en/apartment-rebecca-p1954',
             'https://www.visitopatija.com/en/villa-de-zamagna-apartment-angiolina-p1882',
             'https://www.visitopatija.com/en/villa-penkala-p1948',
             'https://www.visitopatija.com/en/apartments-rumac-p1875',
             'https://www.visitopatija.com/en/ascic-gudeljevic-anica-apartman-3-p1000',
             'https://www.visitopatija.com/en/design-apartment-opatija-p1580',
             'https://www.visitopatija.com/en/apartments-sutalo-p1612',
             'https://www.visitopatija.com/en/apartment-casa-mandarin-p1962',
             'https://www.visitopatija.com/en/apartament-mediteraneo-p1866',
             'https://www.visitopatija.com/en/apartment-bucic-milovan-1-p1058',
             'https://www.visitopatija.com/en/apartment-envi-p1870',
             'https://www.visitopatija.com/en/seaside-apartment-volosko-p2048',
             'https://www.visitopatija.com/en/apartment-zeko-p1815',
             'https://www.visitopatija.com/en/studio-apartments-vila-sokol-p1504',
             'https://www.visitopatija.com/en/apartment-izabela-p2088',
             'https://www.visitopatija.com/en/apartment-ruzica-p2075',
             'https://www.visitopatija.com/en/studio-apartment-tomic-2-p922',
             'https://www.visitopatija.com/en/studio-apartment-ascic-gudeljevic-p965',
             'https://www.visitopatija.com/en/apartments-anita-p967',
             'https://www.visitopatija.com/en/studio-apartment-biba-p1026',
             'https://www.visitopatija.com/en/apartments-juraga-villa-salone-p1110',
             'https://www.visitopatija.com/en/villa-dinka-apartment-p1114',
             'https://www.visitopatija.com/en/studio-apartment-hana-p1209',
             'https://www.visitopatija.com/en/studio-apartment-gudeljevic-p1265',
             'https://www.visitopatija.com/en/peaceful-apartment-in-opatija-p1363',
             'https://www.visitopatija.com/en/apartments-vivoda-lina-p1377',
             'https://www.visitopatija.com/en/studio-apartment-vivoda-p1382',
             'https://www.visitopatija.com/en/mahacek-franjo-p1396',
             'https://www.visitopatija.com/en/studio-apartment-luna-p1441',
             'https://www.visitopatija.com/en/mirko-bulic-p1495',
             'https://www.visitopatija.com/en/studio-apartments-vila-sokol-p1504',
             'https://www.visitopatija.com/en/studio-apartment-nada-volosko-p1511',
             'https://www.visitopatija.com/en/studio-apartment-ramadan-iriskic-p1590',
             'https://www.visitopatija.com/en/studio-apartment-slavijana-trinajstic-p1703',
             'https://www.visitopatija.com/en/studio-apartment-vilma-p1773',
             'https://www.visitopatija.com/en/apartmans-vladislav-vivoda-p1780',
             'https://www.visitopatija.com/en/studio-apartment-zeko-p1816',
             'https://www.visitopatija.com/en/studio-apartment-ujcic-p1840',
             'https://www.visitopatija.com/en/studio-apartment-mirko-djolonga-1-p1841',
             'https://www.visitopatija.com/en/studio-apartment-mirko-djolonga-2-p1842',
             'https://www.visitopatija.com/en/studio-apartment-mirko-djolonga-3-p1843',
             'https://www.visitopatija.com/en/apartments-agata-p1848',
             'https://www.visitopatija.com/en/studio-apartment-ivela-p1852',
             'https://www.visitopatija.com/en/studio-apartment-amorino-p1858',
             'https://www.visitopatija.com/en/villa-antonija-p1890',
             'https://www.visitopatija.com/en/studio-apartments-lauro-p1898',
             'https://www.visitopatija.com/en/villa-klaudia-p1949',
             'https://www.visitopatija.com/en/studio-apartment-teo-p1952',
             'https://www.visitopatija.com/en/studio-apartman-lusetic-p1953',
             'https://www.visitopatija.com/en/villa-salus-apartment-maruna-p1959',
             'https://www.visitopatija.com/en/vila-salus-studio-apartman-buchholz-p1963',
             'https://www.visitopatija.com/en/vila-ozon-p1982',
             'https://www.visitopatija.com/en/studio-apartment-petit-val-p2035',
             'https://www.visitopatija.com/en/-p2079', 'https://www.visitopatija.com/en/studio-rozi-p2083',
             'https://www.visitopatija.com/en/-p2086', 'https://www.visitopatija.com/en/room-arnaldo-malnig-p997',
             'https://www.visitopatija.com/en/apartment-opatija-lipovicaborna-cuk-p1043',
             'https://www.visitopatija.com/en/rooms-dina-iriskic-p1113',
             'https://www.visitopatija.com/en/dragicevic-darjan-soba-p1125',
             'https://www.visitopatija.com/en/rooms-ujcic-p1156', 'https://www.visitopatija.com/en/villa-ika-p1304',
             'https://www.visitopatija.com/en/room-iris-p1850', 'https://www.visitopatija.com/en/room-ivela-p1853',
             'https://www.visitopatija.com/en/villa-klaudia-p1950',
             'https://www.visitopatija.com/en/vacation-home-opatija-p917',
             'https://www.visitopatija.com/en/vacation-house-domizil-p1217',
             'https://www.visitopatija.com/en/villa-marina-p1437', 'https://www.visitopatija.com/en/villa-sole-p1446',
             'https://www.visitopatija.com/en/villa-marone-p1845',
             'https://www.visitopatija.com/en/vacation-house-lidija-p1863',
             'https://www.visitopatija.com/en/villa-palma-p1892',
             'https://www.visitopatija.com/en/healty-house-holiday-home-p1897',
             'https://www.visitopatija.com/en/villa-visum-p1901',
             'https://www.visitopatija.com/en/villa-natura-2-p1902',
             'https://www.visitopatija.com/en/villa-natura1-p1903',
             'https://www.visitopatija.com/en/villa-natura-3-p1904',
             'https://www.visitopatija.com/en/villa-natura-4-p1905',
             'https://www.visitopatija.com/en/villa-natura-5-p1906',
             'https://www.visitopatija.com/en/villa-sv-josip-p1940',
             'https://www.visitopatija.com/en/villa-sv-theodor-p1941',
             'https://www.visitopatija.com/en/villa-veprina-p2041',
             'https://www.visitopatija.com/en/-p2074']


class Visto:
    def __init__(self, teardown=True):

        s = Service(ChromeDriverManager().install())
        self.options = webdriver.ChromeOptions()
        self.options.add_argument('headless')
        self.teardown = teardown
        # keep chrome open
        # self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option(
            "excludeSwitches",
            ['enable-logging'])
        self.driver = webdriver.Chrome(
            options=self.options,
            service=s)
        self.driver.implicitly_wait(50)
        self.category_link = [
            "https://www.visitopatija.com/en/apartments-t195",
            "https://www.visitopatija.com/en/studio-apartments-t196",
            "https://www.visitopatija.com/en/rooms-t198",
            "https://www.visitopatija.com/en/vacation-houses-t197",
        ]
        self.detail_links = ALL_LINKS
        self.collection = []
        super(Visto, self).__init__()

    def __enter__(self):
        self.driver.get(BASE_URL)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.driver.quit()

    def land_first_page(self):
        self.driver.get(BASE_URL)

    def get_open_all_category_links(self):
        for item in self.category_link:
            self.driver.get(item)

            self.get_all_detail_links()

    def get_all_detail_links(self):
        detail_elements = self.driver.find_elements(
            by=By.TAG_NAME, value="article"
        )
        for detail_element in detail_elements:
            detail_urls = detail_element.find_element(by=By.TAG_NAME, value="a").get_attribute("href").strip()
            self.detail_links.append(detail_urls)

    def get_details(self):
        for item in self.detail_links:
            self.driver.get(item)
            try:
                email = self.driver.find_element(
                    by=By.CSS_SELECTOR, value='li[class="mail"]').find_element(
                    by=By.TAG_NAME, value="a").get_attribute(
                    "href").strip()
                if email:
                    email = email.replace("mailto:", "")
            except:
                email = None

            self.collection.append(
                {
                    "email": email,
                }
            )
            print(email)
            with open('visto.json', 'w') as f:
                f.write(json.dumps(self.collection))

    def convert_to_xl(self):
        file = open('visto.json', 'r')
        read_file = file.read()
        collection = json.loads(read_file)
        # Create a new Excel file
        workbook = openpyxl.Workbook()

        # Get the active sheet
        worksheet = workbook.active

        # Add the headers to the sheet
        headers = list(collection[0].keys())  # Get the keys of the first dictionary in the list
        for col, header in enumerate(headers):
            worksheet.cell(row=1, column=col + 1).value = header

        # Add the collection to the sheet
        for row, item in enumerate(collection):
            for col, value in enumerate(item.values()):
                worksheet.cell(row=row + 2, column=col + 1).value = value

        # Save the file
        workbook.save("visto.xlsx")
