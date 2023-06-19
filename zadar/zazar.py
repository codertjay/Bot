import html
import json

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl

BASE_URL = "https://zadar.travel/accommodation/apartments/"
from time import sleep

ALL_LINKS = [
    "https://zadar.travel/accommodation/apartments/irena-roncevic",
    "https://zadar.travel/accommodation/apartments/sky-sun-rooms",
    "https://zadar.travel/accommodation/apartments/benak-marko",
    "https://zadar.travel/accommodation/apartments/guest-house-pegla",
    "https://zadar.travel/accommodation/apartments/lux",
    "https://zadar.travel/accommodation/apartments/apartment-luka-arbanasi",
    "https://zadar.travel/accommodation/apartments/lucic-apartments",
    "https://zadar.travel/accommodation/apartments/bagi",
    "https://zadar.travel/accommodation/apartments/sea-organ-jokic-dragomir",
    "https://zadar.travel/accommodation/apartments/studio-rb",
    "https://zadar.travel/accommodation/apartments/studio-enio-ii",
    "https://zadar.travel/accommodation/apartments/studio-apartment-av",
    "https://zadar.travel/accommodation/apartments/arsenal-accommodation",
    "https://zadar.travel/accommodation/apartments/buljan",
    "https://zadar.travel/accommodation/apartments/apartment-dominik",
    "https://zadar.travel/accommodation/apartments/feljan-zeljko",
    "https://zadar.travel/accommodation/apartments/sara",
    "https://zadar.travel/accommodation/apartments/sara-ii",
    "https://zadar.travel/accommodation/apartments/studio-lucia-zadar-old-town",
    "https://zadar.travel/accommodation/apartments/studio-mare",
    "https://zadar.travel/accommodation/apartments/villa-ibis",
    "https://zadar.travel/accommodation/apartments/zhouse-inn",
    "https://zadar.travel/accommodation/apartments/apartment-ravkin",
    "https://zadar.travel/accommodation/apartments/sea-house-rava",
    "https://zadar.travel/accommodation/apartments/studio-enio-i",
    "https://zadar.travel/accommodation/apartments/apartments-stana",
    "https://zadar.travel/accommodation/apartments/rustica-zadar-with-a-private-pool",
    "https://zadar.travel/accommodation/apartments/apartment-miodrag-ii",
    "https://zadar.travel/accommodation/apartments/apartment-casali",
    "https://zadar.travel/accommodation/apartments/apartment-tina",
    "https://zadar.travel/accommodation/apartments/apartment-green-oasis",
    "https://zadar.travel/accommodation/apartments/villa-masline",
    "https://zadar.travel/accommodation/apartments/apartment-maya-305",
    "https://zadar.travel/accommodation/apartments/apartment-tomas",
    "https://zadar.travel/accommodation/apartments/apartment-anina",
    "https://zadar.travel/accommodation/apartments/apartman-luce",
    "https://zadar.travel/accommodation/apartments/apartments-tonci",
    "https://zadar.travel/accommodation/apartments/apartment-paula",
    "https://zadar.travel/accommodation/apartments/sunset-holiday-house",
    "https://zadar.travel/accommodation/apartments/cosy-family-friendly-house",
    "https://zadar.travel/accommodation/apartments/marija-apartment",
    "https://zadar.travel/accommodation/apartments/city-center-apartment-zadar",
    "https://zadar.travel/accommodation/apartments/milanja-aldo",
    "https://zadar.travel/accommodation/apartments/apartman-antonia",
    "https://zadar.travel/accommodation/apartments/apartment-lara",
    "https://zadar.travel/accommodation/apartments/studio-apartment-sirena",
    "https://zadar.travel/accommodation/apartments/apartments-marija",
    "https://zadar.travel/accommodation/apartments/apartment-cherry",
    "https://zadar.travel/accommodation/apartments/apartments-mike",
    "https://zadar.travel/accommodation/apartments/paola-apartment",
    "https://zadar.travel/accommodation/apartments/apartment-papa",
    "https://zadar.travel/accommodation/apartments/natalie-apartment",
    "https://zadar.travel/accommodation/apartments/dragica-apartment",
    "https://zadar.travel/accommodation/apartments/studio-apartment-jacqueline",
    "https://zadar.travel/accommodation/apartments/absolute-center-grabovac-katarina",
    "https://zadar.travel/accommodation/apartments/diklo-beach",
    "https://zadar.travel/accommodation/apartments/the-little-prince",
    "https://zadar.travel/accommodation/apartments/residence-25",
    "https://zadar.travel/accommodation/apartments/d-rooms",
    "https://zadar.travel/accommodation/apartments/zen-luxury-apartments",
    "https://zadar.travel/accommodation/apartments/apartments-sinjorina",
    "https://zadar.travel/accommodation/apartments/apartments-maras-i",
    "https://zadar.travel/accommodation/apartments/calm-accommodation",
    "https://zadar.travel/accommodation/apartments/apartments-gardelin",
    "https://zadar.travel/accommodation/apartments/apartment-nela",
    "https://zadar.travel/accommodation/apartments/deluxe-apartments-near-the-bridge",
    "https://zadar.travel/accommodation/apartments/dalti-center-apartment",
    "https://zadar.travel/accommodation/apartments/unique-luxury-designers-apartment",
    "https://zadar.travel/accommodation/apartments/apartments-luky-zadar",
    "https://zadar.travel/accommodation/apartments/sens-apartments",
    "https://zadar.travel/accommodation/apartments/hobbit-1",
    "https://zadar.travel/accommodation/apartments/contessa-vitali-luxury-rooms",
    "https://zadar.travel/accommodation/apartments/djana-lucic-matija",
    "https://zadar.travel/accommodation/apartments/ema-miro-tolic",
    "https://zadar.travel/accommodation/apartments/fanganel",
    "https://zadar.travel/accommodation/apartments/goga",
    "https://zadar.travel/accommodation/apartments/hegic",
    "https://zadar.travel/accommodation/apartments/ina-slavko-tresoglavic",
    "https://zadar.travel/accommodation/apartments/ivana-franov-jadranka",
    "https://zadar.travel/accommodation/apartments/jukic",
    "https://zadar.travel/accommodation/apartments/mira",
    "https://zadar.travel/accommodation/apartments/pavle-grgurev-ivana",
    "https://zadar.travel/accommodation/apartments/zvonkoperica",
    "https://zadar.travel/accommodation/apartments/soul-of-dalmatia-baljak-tanja",
    "https://zadar.travel/accommodation/apartments/studio-gloria-accatino-paolo",
    "https://zadar.travel/accommodation/apartments/tim-apartments",
    "https://zadar.travel/accommodation/apartments/villa-clementa",
    "https://zadar.travel/accommodation/apartments/villa-ivana",
    "https://zadar.travel/accommodation/apartments/apartments-zrilic",
    "https://zadar.travel/accommodation/apartments/dragan-zuban",
    "https://zadar.travel/accommodation/apartments/vojmir-loncar",
    "https://zadar.travel/accommodation/apartments/marcel-zadar-city-center",
    "https://zadar.travel/accommodation/apartments/calle-larga-apartment-in-the-old-town",
    "https://zadar.travel/accommodation/apartments/blazenka-zadar-borik-jere-pavicic",
    "https://zadar.travel/accommodation/apartments/rooms-tisa-old-town",
    "https://zadar.travel/accommodation/apartments/zadar-street-apartments",
    "https://zadar.travel/accommodation/apartments/maremy",
    "https://zadar.travel/accommodation/apartments/apartment-jm",
    "https://zadar.travel/accommodation/apartments/apartment-veronika",
    "https://zadar.travel/accommodation/apartments/apartment-matela",
    "https://zadar.travel/accommodation/apartments/apartments-manda",
    "https://zadar.travel/accommodation/apartments/apartmani-simicev",
    "https://zadar.travel/accommodation/apartments/apartments-onorina",
    "https://zadar.travel/accommodation/apartments/beach-paradise-apartments",
    "https://zadar.travel/accommodation/apartments/pavles-apartments",
    "https://zadar.travel/accommodation/apartments/ivy-apartments-zadar",
    "https://zadar.travel/accommodation/apartments/villa-micic",
    "https://zadar.travel/accommodation/apartments/globe-apartament-2",
    "https://zadar.travel/accommodation/apartments/apartment-delfina",
    "https://zadar.travel/accommodation/apartments/studio-libra-zadar",
    "https://zadar.travel/accommodation/apartments/apartment-nekic",
    "https://zadar.travel/accommodation/apartments/apartments-maja",
    "https://zadar.travel/accommodation/apartments/old-town-sea-organ-studio-apartment",
    "https://zadar.travel/accommodation/apartments/apartman-sweet-odette",
    "https://zadar.travel/accommodation/apartments/apartments-da",
    "https://zadar.travel/accommodation/apartments/apartments-marea",
    "https://zadar.travel/accommodation/apartments/kailani-luxury-apartment-room",
    "https://zadar.travel/accommodation/apartments/apartments-adriatic",
    "https://zadar.travel/accommodation/apartments/ap-silvija",
    "https://zadar.travel/accommodation/apartments/villa-orion",
    "https://zadar.travel/accommodation/apartments/family-apartments-zadar",
    "https://zadar.travel/accommodation/apartments/griva-apartments",
    "https://zadar.travel/accommodation/apartments/jeta-studio-rooms",
    "https://zadar.travel/accommodation/apartments/villa-persic",
    "https://zadar.travel/accommodation/apartments/tereza-apartments",
    "https://zadar.travel/accommodation/apartments/apartment-star",
    "https://zadar.travel/accommodation/apartments/apartment-toni",
    "https://zadar.travel/accommodation/apartments/sound-of-the-sea",
    "https://zadar.travel/accommodation/apartments/apartment-lavanda",
    "https://zadar.travel/accommodation/apartments/apartments-talir",
    "https://zadar.travel/accommodation/apartments/apartment-milin",
    "https://zadar.travel/accommodation/apartments/apartments-santini-my-dream",
    "https://zadar.travel/accommodation/apartments/apartment-beti",
    "https://zadar.travel/accommodation/apartments/peninsula-rooms",
    "https://zadar.travel/accommodation/apartments/apartments-mario-pustahija",
    "https://zadar.travel/accommodation/apartments/guesthouse-villa-irena",
    "https://zadar.travel/accommodation/apartments/the-2sisters-suite",
    "https://zadar.travel/accommodation/apartments/apartments-roncevic",
    "https://zadar.travel/accommodation/apartments/apartments-marko",
    "https://zadar.travel/accommodation/apartments/mb-luxury-apartment-1",
    "https://zadar.travel/accommodation/apartments/apartments-mato",
    "https://zadar.travel/accommodation/apartments/legacy-marine-1-luxury-family-penthouse",
    "https://zadar.travel/accommodation/apartments/ruzmarin",
    "https://zadar.travel/accommodation/apartments/lavanda-zadar",
    "https://zadar.travel/accommodation/apartments/apartments-kekic",
    "https://zadar.travel/accommodation/apartments/villa-nada",
    "https://zadar.travel/accommodation/apartments/miriam-apartments",
    "https://zadar.travel/accommodation/apartments/apartment-adriana",
    "https://zadar.travel/accommodation/apartments/mariola-apartment",
    "https://zadar.travel/accommodation/apartments/apartments-luky-2-zadar",
    "https://zadar.travel/accommodation/apartments/apartment-luka-molat",
    "https://zadar.travel/accommodation/apartments/apartments-aurora",
    "https://zadar.travel/accommodation/apartments/apartment-m",
    "https://zadar.travel/accommodation/apartments/serenity-apartment",
    "https://zadar.travel/accommodation/apartments/apartment-porthos",
    "https://zadar.travel/accommodation/apartments/apartments-hana-zadar",
    "https://zadar.travel/accommodation/apartments/apartments-blue",
    "https://zadar.travel/accommodation/apartments/denis-apartments",
    "https://zadar.travel/accommodation/apartments/anastazia-apartment",
    "https://zadar.travel/accommodation/apartments/lino-apartment",
    "https://zadar.travel/accommodation/apartments/lily-apartment",
    "https://zadar.travel/accommodation/apartments/emma-apartment",
    "https://zadar.travel/accommodation/apartments/villa-brucker",
    "https://zadar.travel/accommodation/apartments/lovre-apartment",
    "https://zadar.travel/accommodation/apartments/galija-apartment",
    "https://zadar.travel/accommodation/apartments/silvia-apartments3",
    "https://zadar.travel/accommodation/apartments/matea-apartments",
    "https://zadar.travel/accommodation/apartments/nevena-studio-apartment",
    "https://zadar.travel/accommodation/apartments/apartment-in-zadar",
    "https://zadar.travel/accommodation/apartments/yellow-cat-apartment",
    "https://zadar.travel/accommodation/apartments/seabreeze-apartment",
    "https://zadar.travel/accommodation/apartments/rede-apartments-in-diklo",
    "https://zadar.travel/accommodation/apartments/kesina-apartments",
    "https://zadar.travel/accommodation/apartments/blue-sea-apartment",
    "https://zadar.travel/accommodation/apartments/ela-apartment",
    "https://zadar.travel/accommodation/apartments/apartment-pegy",
    "https://zadar.travel/accommodation/apartments/nena-apartment",
    "https://zadar.travel/accommodation/apartments/guest-house-baric",
    "https://zadar.travel/accommodation/apartments/apartment-m1",
    "https://zadar.travel/accommodation/apartments/davorka-vacation-house",
    "https://zadar.travel/accommodation/apartments/mate-galleria-apartment",
    "https://zadar.travel/accommodation/apartments/apartments-zara",
    "https://zadar.travel/accommodation/apartments/cosy-villa-near-the-beach",
    "https://zadar.travel/accommodation/apartments/rava-art-amphora-studioroom",
    "https://zadar.travel/accommodation/apartments/apartment-bliss",
    "https://zadar.travel/accommodation/apartments/apartment-mille-stelle",
    "https://zadar.travel/accommodation/apartments/apartment-mira",
    "https://zadar.travel/accommodation/apartments/apartment-cecilija",
    "https://zadar.travel/accommodation/apartments/apartments-aurora-zadar",
    "https://zadar.travel/accommodation/apartments/minipalais",
    "https://zadar.travel/accommodation/apartments/ante",
    "https://zadar.travel/accommodation/apartments/edi-1",
    "https://zadar.travel/accommodation/apartments/zrinka-matulic-djino",
    "https://zadar.travel/accommodation/apartments/zuvela-damir",
    "https://zadar.travel/accommodation/apartments/apartment-franica",
    "https://zadar.travel/accommodation/apartments/rooms-nekic-1",
    "https://zadar.travel/accommodation/apartments/rooms-nekic",
    "https://zadar.travel/accommodation/apartments/apartment-marija-kozino",
    "https://zadar.travel/accommodation/apartments/apartment-rafaela",
    "https://zadar.travel/accommodation/apartments/apartments-nada",
    "https://zadar.travel/accommodation/apartments/apartments-kristo-zadar",
    "https://zadar.travel/accommodation/apartments/apartment-karla",
    "https://zadar.travel/accommodation/apartments/denis-apartments1",
    "https://zadar.travel/accommodation/apartments/sanja-studio-apartments",
    "https://zadar.travel/accommodation/apartments/apartment-nora",
    "https://zadar.travel/accommodation/apartments/apartment-fontana-felicita",
    "https://zadar.travel/accommodation/apartments/apartment-lavanda-home",
    "https://zadar.travel/accommodation/apartments/studio-apartment-tristan",
    "https://zadar.travel/accommodation/apartments/apartment-xenia",
    "https://zadar.travel/accommodation/apartments/apartment-ika",
    "https://zadar.travel/accommodation/apartments/wave-apartment",
    "https://zadar.travel/accommodation/apartments/roto",
    "https://zadar.travel/accommodation/apartments/apartment-anita",
    "https://zadar.travel/accommodation/apartments/apartments-birgit",
    "https://zadar.travel/accommodation/apartments/diva",
    "https://zadar.travel/accommodation/apartments/padrenostro",
    "https://zadar.travel/accommodation/apartments/punta-skala-43",
    "https://zadar.travel/accommodation/apartments/secret-garden-apartment",
    "https://zadar.travel/accommodation/apartments/high-street-modern-apartment",
    "https://zadar.travel/accommodation/apartments/apartment-thea",
    "https://zadar.travel/accommodation/apartments/piazza-rooms-zadar-old-town",
    "https://zadar.travel/accommodation/apartments/little-gem-in-the-city-center",
    "https://zadar.travel/accommodation/apartments/bike-corner",
    "https://zadar.travel/accommodation/apartments/apartment-ivana",
    "https://zadar.travel/accommodation/apartments/apartman-jakov",
    "https://zadar.travel/accommodation/apartments/luxury-apartment-cityzen",
    "https://zadar.travel/accommodation/apartments/diva-studio",
    "https://zadar.travel/accommodation/apartments/villa-vigo-apartment-with-a-pool",
    "https://zadar.travel/accommodation/apartments/studio-luka",
    "https://zadar.travel/accommodation/apartments/rusula",
    "https://zadar.travel/accommodation/apartments/zara-lumina-apartment",
    "https://zadar.travel/accommodation/apartments/snijezana-genari",
    "https://zadar.travel/accommodation/apartments/studio-apartment-luna",
    "https://zadar.travel/accommodation/apartments/apartment-luna",
    "https://zadar.travel/accommodation/apartments/apartman-borik-terrace-sea-view",
    "https://zadar.travel/accommodation/apartments/apartment-alegria-zadar",
    "https://zadar.travel/accommodation/apartments/apartment-anita-petrcane",
    "https://zadar.travel/accommodation/apartments/tajana-polic-studio-apartments",
    "https://zadar.travel/accommodation/apartments/apartment-rubica",
    "https://zadar.travel/accommodation/apartments/apartments-sanja",
    "https://zadar.travel/accommodation/apartments/petrcane-apartment-maslina",
    "https://zadar.travel/accommodation/apartments/borgo",
    "https://zadar.travel/accommodation/apartments/goodbed-zadar",
    "https://zadar.travel/accommodation/apartments/holiday-center-ivona",
    "https://zadar.travel/accommodation/apartments/lily-of-the-valley",
    "https://zadar.travel/accommodation/apartments/niko",
    "https://zadar.travel/accommodation/apartments/opera",
    "https://zadar.travel/accommodation/apartments/teuta",
    "https://zadar.travel/accommodation/apartments/villa-vanilla",
    "https://zadar.travel/accommodation/apartments/villa-vesna",
    "https://zadar.travel/accommodation/apartments/villa-violet",
    "https://zadar.travel/accommodation/apartments/zadar-city-studios",
    "https://zadar.travel/accommodation/apartments/zadractive",
    "https://zadar.travel/accommodation/apartments/gorana-medic",
    "https://zadar.travel/accommodation/apartments/andrija-jovic",
    "https://zadar.travel/accommodation/apartments/lipo-je",
    "https://zadar.travel/accommodation/apartments/centrooms-kovac",
    "https://zadar.travel/accommodation/apartments/tammy-apartment",
    "https://zadar.travel/accommodation/apartments/villa-sportiva-zadar",
    "https://zadar.travel/accommodation/apartments/apartment-dioniz",
    "https://zadar.travel/accommodation/apartments/luxury-apartment-luna-with-a-pool",
    "https://zadar.travel/accommodation/apartments/classy-design-accommodation",
    "https://zadar.travel/accommodation/apartments/villa-stella-four-star-apartment",
    "https://zadar.travel/accommodation/apartments/apriori-apartments",
    "https://zadar.travel/accommodation/apartments/ada",
    "https://zadar.travel/accommodation/apartments/alex",
    "https://zadar.travel/accommodation/apartments/amelie",
    "https://zadar.travel/accommodation/apartments/mirjan-radman",
    "https://zadar.travel/accommodation/apartments/andria",
    "https://zadar.travel/accommodation/apartments/andela",
    "https://zadar.travel/accommodation/apartments/andelo",
    "https://zadar.travel/accommodation/apartments/ankica",
    "https://zadar.travel/accommodation/apartments/arsenal-1",
    "https://zadar.travel/accommodation/apartments/bokanjac-i",
    "https://zadar.travel/accommodation/apartments/bozena",
    "https://zadar.travel/accommodation/apartments/branimir",
    "https://zadar.travel/accommodation/apartments/bruno",
    "https://zadar.travel/accommodation/apartments/bukvic",
    "https://zadar.travel/accommodation/apartments/centar-zara",
    "https://zadar.travel/accommodation/apartments/chicquiet",
    "https://zadar.travel/accommodation/apartments/darija",
    "https://zadar.travel/accommodation/apartments/diklo",
    "https://zadar.travel/accommodation/apartments/dupin",
    "https://zadar.travel/accommodation/apartments/el-mirador",
    "https://zadar.travel/accommodation/apartments/ero",
    "https://zadar.travel/accommodation/apartments/eva",
    "https://zadar.travel/accommodation/apartments/forum",
    "https://zadar.travel/accommodation/apartments/franka",
    "https://zadar.travel/accommodation/apartments/genari",
    "https://zadar.travel/accommodation/apartments/giulia",
    "https://zadar.travel/accommodation/apartments/golden-gate-zadar",
    "https://zadar.travel/accommodation/apartments/hajnc",
    "https://zadar.travel/accommodation/apartments/heart-of-city-lola",
    "https://zadar.travel/accommodation/apartments/idassa",
    "https://zadar.travel/accommodation/apartments/idassa-palace",
    "https://zadar.travel/accommodation/apartments/ivka-1",
    "https://zadar.travel/accommodation/apartments/ivka-2",
    "https://zadar.travel/accommodation/apartments/ivo",
    "https://zadar.travel/accommodation/apartments/ivona",
    "https://zadar.travel/accommodation/apartments/josipa",
    "https://zadar.travel/accommodation/apartments/jovic",
    "https://zadar.travel/accommodation/apartments/juraj",
    "https://zadar.travel/accommodation/apartments/kala",
    "https://zadar.travel/accommodation/apartments/karla",
    "https://zadar.travel/accommodation/apartments/karmen",
    "https://zadar.travel/accommodation/apartments/katia",
    "https://zadar.travel/accommodation/apartments/krijanca",
    "https://zadar.travel/accommodation/apartments/la-la",
    "https://zadar.travel/accommodation/apartments/lavanda",
    "https://zadar.travel/accommodation/apartments/lea",
    "https://zadar.travel/accommodation/apartments/lile",
    "https://zadar.travel/accommodation/apartments/linda-mario",
    "https://zadar.travel/accommodation/apartments/ljubica",
    "https://zadar.travel/accommodation/apartments/lovely-maria",
    "https://zadar.travel/accommodation/apartments/luce-mala",
    "https://zadar.travel/accommodation/apartments/mara",
    "https://zadar.travel/accommodation/apartments/marea",
    "https://zadar.travel/accommodation/apartments/marina",
    "https://zadar.travel/accommodation/apartments/marty",
    "https://zadar.travel/accommodation/apartments/medun",
    "https://zadar.travel/accommodation/apartments/miranda",
    "https://zadar.travel/accommodation/apartments/miriam",
    "https://zadar.travel/accommodation/apartments/mirjana",
    "https://zadar.travel/accommodation/apartments/mirta-s-loft",
    "https://zadar.travel/accommodation/apartments/molly",
    "https://zadar.travel/accommodation/apartments/morovic",
    "https://zadar.travel/accommodation/apartments/nela",
    "https://zadar.travel/accommodation/apartments/nena",
    "https://zadar.travel/accommodation/apartments/nika-apartments",
    "https://zadar.travel/accommodation/apartments/nita",
    "https://zadar.travel/accommodation/apartments/old-cottage",
    "https://zadar.travel/accommodation/apartments/penthouse-petrcani",
    "https://zadar.travel/accommodation/apartments/pero",
    "https://zadar.travel/accommodation/apartments/popona",
    "https://zadar.travel/accommodation/apartments/premuda",
    "https://zadar.travel/accommodation/apartments/punta-skala",
    "https://zadar.travel/accommodation/apartments/rafael",
    "https://zadar.travel/accommodation/apartments/robert",
    "https://zadar.travel/accommodation/apartments/rosa",
    "https://zadar.travel/accommodation/apartments/sea-organ-zadar",
    "https://zadar.travel/accommodation/apartments/silvija",
    "https://zadar.travel/accommodation/apartments/simeon-suite",
    "https://zadar.travel/accommodation/apartments/skalinada",
    "https://zadar.travel/accommodation/apartments/stanko",
    "https://zadar.travel/accommodation/apartments/studio",
    "https://zadar.travel/accommodation/apartments/sunset-red",
    "https://zadar.travel/accommodation/apartments/suzi",
    "https://zadar.travel/accommodation/apartments/sweet-dreams",
    "https://zadar.travel/accommodation/apartments/tea",
    "https://zadar.travel/accommodation/apartments/the-flat",
    "https://zadar.travel/accommodation/apartments/victoria",
    "https://zadar.travel/accommodation/apartments/victoria-ii",
    "https://zadar.travel/accommodation/apartments/victoria-iii",
    "https://zadar.travel/accommodation/apartments/villa-amantina",
    "https://zadar.travel/accommodation/apartments/villa-damir",
    "https://zadar.travel/accommodation/apartments/villa-katarina",
    "https://zadar.travel/accommodation/apartments/villa-mandolina",
    "https://zadar.travel/accommodation/apartments/villa-marija",
    "https://zadar.travel/accommodation/apartments/villa-mirella",
    "https://zadar.travel/accommodation/apartments/studio-nodi",
    "https://zadar.travel/accommodation/apartments/villa-predovan",
    "https://zadar.travel/accommodation/apartments/villa-triana",
    "https://zadar.travel/accommodation/apartments/villa-valentina",
    "https://zadar.travel/accommodation/apartments/zadar-4-you",
    "https://zadar.travel/accommodation/apartments/zan-i-anamarija",
    "https://zadar.travel/accommodation/apartments/zara",
    "https://zadar.travel/accommodation/apartments/zara-city-apartments",
    "https://zadar.travel/accommodation/apartments/larisa-pavlovic-petani",
    "https://zadar.travel/accommodation/apartments/vesna",
    "https://zadar.travel/accommodation/apartments/boris-gacina",
    "https://zadar.travel/accommodation/apartments/jurjevic-apartment-punta-skala",
    "https://zadar.travel/accommodation/apartments/davor-sambunjak",
    "https://zadar.travel/accommodation/apartments/martha-deluxe",
    "https://zadar.travel/accommodation/apartments/apartments-lenko",
    "https://zadar.travel/accommodation/apartments/apartment-palatisa",
    "https://zadar.travel/accommodation/apartments/sunny-place",
    "https://zadar.travel/accommodation/apartments/apartment-2-relax-zadar",
    "https://zadar.travel/accommodation/apartments/em-room-zadar",
    "https://zadar.travel/accommodation/apartments/latmo-apartment-zadar",
    "https://zadar.travel/accommodation/apartments/apartment-nada-albert-fatovic",
    "https://zadar.travel/accommodation/apartments/apartments-summer-touch-petrcane",
    "https://zadar.travel/accommodation/apartments/apartments-valentina",
    "https://zadar.travel/accommodation/apartments/apartment-andreja",
    "https://zadar.travel/accommodation/apartments/studio-apartment-dragica",
    "https://zadar.travel/accommodation/apartments/apartments-ema-zadar",
    "https://zadar.travel/accommodation/apartments/apartments-dragic",
    "https://zadar.travel/accommodation/apartments/apartment-petkovic",
    "https://zadar.travel/accommodation/apartments/new-luxurious-apartment-loreta",
    "https://zadar.travel/accommodation/apartments/apartments-rotim",
    "https://zadar.travel/accommodation/apartments/apartment-meri",
    "https://zadar.travel/accommodation/apartments/apartment-kister",
    "https://zadar.travel/accommodation/apartments/andreas-apartment",
    "https://zadar.travel/accommodation/apartments/sun-and-sea-apartment",
    "https://zadar.travel/accommodation/apartments/sunshine-near-bridge-apartment",
    "https://zadar.travel/accommodation/apartments/villa-marijana-kozino",
    "https://zadar.travel/accommodation/apartments/friendly-lovely-cosy-apartment",
    "https://zadar.travel/accommodation/apartments/skunca-apartment",
    "https://zadar.travel/accommodation/apartments/apartments-spiritus-mare",
    "https://zadar.travel/accommodation/apartments/apartment-near-centre",
    "https://zadar.travel/accommodation/apartments/studio-apartment-fabris",
    "https://zadar.travel/accommodation/apartments/apartment-matea",
    "https://zadar.travel/accommodation/apartments/angela-luxury-apartment-i",
    "https://zadar.travel/accommodation/apartments/musical-paradise-baltazar",
    "https://zadar.travel/accommodation/apartments/villa-stella-three-star-apartments",
    "https://zadar.travel/accommodation/apartments/apartman-binka-punta-skala",
    "https://zadar.travel/accommodation/apartments/dj-apartment",
    "https://zadar.travel/accommodation/apartments/apartment-pavla",
    "https://zadar.travel/accommodation/apartments/studio-maria",
    "https://zadar.travel/accommodation/apartments/apartment-petar",
    "https://zadar.travel/accommodation/apartments/exclusive-residence-apartment",
    "https://zadar.travel/accommodation/apartments/doda-punta-skala-doris-bronic-bilac",
    "https://zadar.travel/accommodation/apartments/donata",
    "https://zadar.travel/accommodation/apartments/downtown",
    "https://zadar.travel/accommodation/apartments/guest-house-marija",
    "https://zadar.travel/accommodation/apartments/apartment-juan",
    "https://zadar.travel/accommodation/apartments/jurisic",
    "https://zadar.travel/accommodation/apartments/katarina",
    "https://zadar.travel/accommodation/apartments/klara",
    "https://zadar.travel/accommodation/apartments/kolega",
    "https://zadar.travel/accommodation/apartments/marijo",
    "https://zadar.travel/accommodation/apartments/apartment-petar",
    "https://zadar.travel/accommodation/apartments/exclusive-residence-apartment",
    "https://zadar.travel/accommodation/apartments/doda-punta-skala-doris-bronic-bilac",
    "https://zadar.travel/accommodation/apartments/donata",
    "https://zadar.travel/accommodation/apartments/downtown",
    "https://zadar.travel/accommodation/apartments/guest-house-marija",
    "https://zadar.travel/accommodation/apartments/apartment-juan",
    "https://zadar.travel/accommodation/apartments/jurisic",
    "https://zadar.travel/accommodation/apartments/katarina",
    "https://zadar.travel/accommodation/apartments/klara",
    "https://zadar.travel/accommodation/apartments/kolega",
    "https://zadar.travel/accommodation/apartments/marijo",
    "https://zadar.travel/accommodation/apartments/marinsko",
    "https://zadar.travel/accommodation/apartments/markov-diklo-ljiljana-markov",
    "https://zadar.travel/accommodation/apartments/mila",
    "https://zadar.travel/accommodation/apartments/natura",
    "https://zadar.travel/accommodation/apartments/paula",
    "https://zadar.travel/accommodation/apartments/paula-petrcane",
    "https://zadar.travel/accommodation/apartments/petra",
    "https://zadar.travel/accommodation/apartments/pulek",
    "https://zadar.travel/accommodation/apartments/rosanda",
    "https://zadar.travel/accommodation/apartments/tara",
    "https://zadar.travel/accommodation/apartments/toka",
    "https://zadar.travel/accommodation/apartments/velimira",
    "https://zadar.travel/accommodation/apartments/victoria-center",
    "https://zadar.travel/accommodation/apartments/doris",
    "https://zadar.travel/accommodation/apartments/city-apartment-marcela",
    "https://zadar.travel/accommodation/apartments/apartments-andela",
    "https://zadar.travel/accommodation/apartments/apartment-busljeta-cosy",
    "https://zadar.travel/accommodation/apartments/petrcane-house",
    "https://zadar.travel/accommodation/apartments/apartment-lucija-marica-nimac",
    "https://zadar.travel/accommodation/apartments/apartment-petra-marica-nimac",
    "https://zadar.travel/accommodation/apartments/apartments-klelija",
    "https://zadar.travel/accommodation/apartments/apartment-robi",
    "https://zadar.travel/accommodation/apartments/zinka-apartment",
    "https://zadar.travel/accommodation/apartments/luka-studio-apartment",

    "https://zadar.travel/accommodation/hotels/falkensteiner-family-hotel-diadora",
    "https://zadar.travel/accommodation/hotels/falkensteiner-hotel-spa-iadera",
    "https://zadar.travel/accommodation/hotels/bastion-heritage-hotel-relais-chateaux",
    "https://zadar.travel/accommodation/hotels/art-hotel-kalelarga",
    "https://zadar.travel/accommodation/hotels/teatro-verdi-boutique-hotel",
    "https://zadar.travel/accommodation/hotels/hotel-kolovare",
    "https://zadar.travel/accommodation/hotels/hotel-petrcane",
    "https://zadar.travel/accommodation/hotels/hotel-pinija",
    "https://zadar.travel/accommodation/hotels/falkensteiner-club-funimation-borik",
    "https://zadar.travel/accommodation/hotels/hotel-mediteran",
    "https://zadar.travel/accommodation/hotels/hotel-donat",
    "https://zadar.travel/accommodation/hotels/hotel-delfin"
]


class Zadar:
    def __init__(self, teardown=True):

        s = Service(ChromeDriverManager().install())
        self.options = webdriver.ChromeOptions()
        self.options.add_argument('headless')
        self.teardown = teardown
        # keep chrome open
        self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option(
            "excludeSwitches",
            ['enable-logging'])
        self.driver = webdriver.Chrome(
            options=self.options,
            service=s)
        self.driver.implicitly_wait(50)
        self.category_link = [
            "https://zadar.travel/accommodation/hotels",
        ]
        self.detail_links = ALL_LINKS
        self.collection = []
        super(Zadar, self).__init__()

    def __enter__(self):
        self.driver.get(BASE_URL)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.driver.quit()

    def land_first_page(self):
        self.driver.get(BASE_URL)

    def get_open_all_category_links(self):
        for item in self.category_link:
            self.driver.get(item)

            self.get_paginated_pages()

    def get_paginated_pages(self):

        buttons_to_click = self.driver.find_elements(
            by=By.CSS_SELECTOR, value='button[class="btn btn--pagination"]'
        )

        for button in buttons_to_click:
            print("Click Button", button.text)

            self.get_links_on_detail_page()
            with open('Zadar_links.json', 'w') as f:
                f.write(json.dumps(self.detail_links))
            try:
                button = self.driver.find_element_by_class_name('btn.btn--pagination')
                self.driver.execute_script("arguments[0].scrollIntoView();", button)
                button.click()
            except:
                print(button.text)
                pass

    def get_links_on_detail_page(self):

        detail_elements = self.driver.find_elements(
            by=By.CSS_SELECTOR, value='a[data-v-0db3bc8b]'
        )
        for detail_element in detail_elements:
            self.detail_links.append(detail_element.get_attribute("href"))

    def get_emails_on_subs(self):
        subs_lists = [
            "https://zadar.travel/accommodation/tourist-agencies",
            "https://zadar.travel/accommodation/marinas-and-yacht-charters/",
            "https://zadar.travel/accommodation/camping",
            "https://zadar.travel/accommodation/hostels",
        ]
        for item in subs_lists:
            self.driver.get(item)
            try:
                elements = self.driver.find_elements(
                    by=By.CSS_SELECTOR, value='a[aria-label="E-mail"]')
                for element in elements:
                    email = element.get_attribute("href").replace("mailto:", "")
                    self.collection.append(
                        {
                            "email": email,
                        }
                    )
                    print(email)
                    with open('ZadarOther.json', 'w') as f:
                        f.write(json.dumps(self.collection))
            except:
                pass

    def get_details(self):
        for item in self.detail_links:
            self.driver.get(item)
            try:
                element = self.driver.find_element(by=By.CSS_SELECTOR,
                                                   value='a[aria-label="Contact e-mail"]').get_attribute("href")
                if element:
                    email = element.replace("mailto:", "")
                    self.collection.append(
                        {
                            "email": email,
                        }
                    )
                    print(email)
                    with open('Zadar.json', 'w') as f:
                        f.write(json.dumps(self.collection))
            except:
                continue

    def convert_to_xl(self):
        file = open('Zadar.json', 'r')
        read_file = file.read()
        collection = json.loads(read_file)
        # Create a new Excel file
        workbook = openpyxl.Workbook()

        # Get the active sheet
        worksheet = workbook.active

        # Add the headers to the sheet
        headers = list(collection[0].keys())  # Get the keys of the first dictionary in the list
        for col, header in enumerate(headers):
            worksheet.cell(row=1, column=col + 1).value = header

        # Add the collection to the sheet
        for row, item in enumerate(collection):
            for col, value in enumerate(item.values()):
                worksheet.cell(row=row + 2, column=col + 1).value = value

        # Save the file
        workbook.save("Zadar.xlsx")